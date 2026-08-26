"""
Migração pontual: bet_tracker_calude.xlsx -> Google Sheets, seguindo o schema
descrito em docs/SCHEMA.md (só dados em bruto; nada de colunas calculadas).

Uso:
    python migrate_to_sheets.py --sheet-id <ID_DA_GOOGLE_SHEET> [--xlsx <caminho>] [--dry-run]

Pré-requisitos (ver migration/README.md para o passo a passo):
    pip install -r requirements.txt
    Uma service account do Google Cloud com a Sheets API ativada, cujo ficheiro
    de credenciais (.json) é apontado pela variável de ambiente
    GOOGLE_APPLICATION_CREDENTIALS. A Google Sheet de destino tem de estar
    partilhada (Editor) com o email dessa service account.

Este script corre uma única vez para semear a Sheet inicial. Depois disso a
Sheet passa a ser a fonte de verdade; este script não faz parte da app.
"""

import argparse
import os
import sys

import openpyxl

DEFAULT_XLSX = r"C:\Users\Samuel\Desktop\bet_tracker_calude.xlsx"

# --- Cabeçalhos por separador (têm de corresponder ao docs/SCHEMA.md) ---

HEADERS = {
    "Apostas": [
        "id_boletim", "data", "tipo_jornada", "jogador", "desporto", "partida",
        "prognostico", "odd", "atraso", "resultado", "observacoes", "pago",
    ],
    "MultaFaltas": ["data", "jogador", "motivo", "estado"],
    "Levantamentos": ["data", "motivo", "valor", "observacoes"],
    "Config_Geral": ["chave", "valor"],
    "Config_Jogadores": ["jogador"],
    "Config_TiposJornada": ["tipo_jornada"],
    "Config_Desportos": ["desporto"],
    "Config_Epocas": ["epoca"],
    "Config_MultaErros": ["nº_errantes", "multa_individual", "multa_total_boletim", "estado_label"],
    "Config_MultaAtrasos": ["nº_atrasos_no_mes", "multa"],
    "Config_MultaFaltas": ["nº_faltas_no_mes", "multa"],
}


def extract(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    apostas_ws = wb["APOSTAS"]
    faltas_ws = wb["MULTA PROGNÓSTICO"]
    lev_ws = wb["LEVANTAMENTOS"]
    cfg_ws = wb["CONFIGURAÇÕES"]

    data = {h: [] for h in HEADERS}

    # Apostas: colunas A,B,C,D,E,F,G,H,I,J,U,V ; linhas 3..max_row
    # id_boletim (A) e tipo_jornada (C) só têm valor na 1ª linha de cada bloco de N pernas
    # no Excel (células fundidas visualmente) - fazemos forward-fill para que cada linha
    # migrada seja autossuficiente (schema em bruto, sem depender de "linha anterior").
    ultimo_id_boletim = None
    ultimo_tipo_jornada = None
    for row in apostas_ws.iter_rows(min_row=3, max_row=apostas_ws.max_row):
        a, b, c, d, e, f, g, h_, i, j = (row[idx].value for idx in range(10))
        u = row[20].value  # coluna U
        v = row[21].value  # coluna V
        if a is None and b is None and d is None:
            continue  # linha vazia
        if a is not None:
            ultimo_id_boletim = a
        if c is not None:
            ultimo_tipo_jornada = c
        data["Apostas"].append([
            ultimo_id_boletim, b.isoformat() if hasattr(b, "isoformat") else b,
            ultimo_tipo_jornada, d, e, f, g, h_, i, j, u or "", v or "",
        ])

    # MultaFaltas: A=data,B=jogador,C=motivo,F=estado ; linhas 3..max_row
    # (separador do .xlsx chama-se "MULTA PROGNÓSTICO", renomeado a partir de "MULTA FALTAS")
    for row in faltas_ws.iter_rows(min_row=3, max_row=faltas_ws.max_row):
        a, b, c = row[0].value, row[1].value, row[2].value
        f_ = row[5].value
        if a is None and b is None:
            continue
        data["MultaFaltas"].append([
            a.isoformat() if hasattr(a, "isoformat") else a, b, c or "", f_ or "",
        ])

    # Levantamentos: A=data,B=motivo,C=valor,D=observacoes ; linhas 3..max_row
    for row in lev_ws.iter_rows(min_row=3, max_row=lev_ws.max_row):
        a, b, c, d = (row[idx].value for idx in range(4))
        if a is None and b is None:
            continue
        data["Levantamentos"].append([
            a.isoformat() if hasattr(a, "isoformat") else a, b, c, d or "",
        ])

    # Config_Geral: chave/valor fixos (localização confirmada por inspeção do .xlsx)
    data["Config_Geral"] = [
        ["banca_inicial", cfg_ws["B4"].value],
        ["montante_por_combinacao", cfg_ws["B5"].value],
        ["lucro_minimo", cfg_ws["B7"].value],
        ["numero_jogadores", cfg_ws["B8"].value],
        ["odd_minima", cfg_ws["B9"].value],
        ["mes_inicio_epoca", 7],  # julho — hoje hard-coded nas fórmulas do Excel (MONTH(...)<7)
    ]

    # Config_Jogadores: A14:A18
    data["Config_Jogadores"] = [[cfg_ws.cell(row=r, column=1).value]
                                 for r in range(14, 19) if cfg_ws.cell(row=r, column=1).value]

    # Config_TiposJornada: A22:A24
    data["Config_TiposJornada"] = [[cfg_ws.cell(row=r, column=1).value]
                                    for r in range(22, 25) if cfg_ws.cell(row=r, column=1).value]

    # Config_Desportos: A28:A33
    data["Config_Desportos"] = [[cfg_ws.cell(row=r, column=1).value]
                                 for r in range(28, 34) if cfg_ws.cell(row=r, column=1).value]

    # Config_Epocas: A37:A41
    data["Config_Epocas"] = [[cfg_ws.cell(row=r, column=1).value]
                              for r in range(37, 42) if cfg_ws.cell(row=r, column=1).value]

    # Config_MultaErros: E5:H10 (nº_errantes, multa_individual, multa_total_boletim, estado_label)
    data["Config_MultaErros"] = [
        [cfg_ws.cell(row=r, column=5).value, cfg_ws.cell(row=r, column=6).value,
         cfg_ws.cell(row=r, column=7).value, cfg_ws.cell(row=r, column=8).value]
        for r in range(5, 11)
    ]

    # Config_MultaAtrasos: E14:F16 (1, 2, 3+)
    data["Config_MultaAtrasos"] = [
        [cfg_ws.cell(row=r, column=5).value, cfg_ws.cell(row=r, column=6).value]
        for r in range(14, 17)
    ]

    # Config_MultaFaltas: E26:F29 (1, 2, 3, 4+)
    data["Config_MultaFaltas"] = [
        [cfg_ws.cell(row=r, column=5).value, cfg_ws.cell(row=r, column=6).value]
        for r in range(26, 30)
    ]

    return data


def push_to_sheets(sheet_id: str, data: dict) -> None:
    import gspread
    from google.oauth2.service_account import Credentials

    creds_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if not creds_path:
        sys.exit(
            "GOOGLE_APPLICATION_CREDENTIALS não está definida. Aponta para o .json "
            "da service account (ver migration/README.md)."
        )

    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)

    for tab_name, header in HEADERS.items():
        rows = data[tab_name]
        try:
            ws = sh.worksheet(tab_name)
            ws.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title=tab_name, rows=max(len(rows) + 10, 20), cols=len(header) + 2)
        ws.update([header] + rows, value_input_option="USER_ENTERED")
        print(f"  {tab_name}: {len(rows)} linha(s) escritas.")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--sheet-id", required=True, help="ID da Google Sheet de destino (da URL)")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Caminho do bet_tracker_calude.xlsx")
    parser.add_argument("--dry-run", action="store_true",
                         help="Só extrai e mostra contagens, não escreve na Sheet")
    args = parser.parse_args()

    print(f"A ler {args.xlsx} ...")
    data = extract(args.xlsx)

    print("\nResumo dos dados extraídos:")
    for tab_name, rows in data.items():
        print(f"  {tab_name}: {len(rows)} linha(s)")

    if args.dry_run:
        print("\n--dry-run: nada foi escrito na Google Sheet.")
        return

    print(f"\nA escrever na Google Sheet {args.sheet_id} ...")
    push_to_sheets(args.sheet_id, data)
    print("\nMigração concluída.")


if __name__ == "__main__":
    main()
